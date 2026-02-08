"""
Report Generator Module for Data Comparison Framework
Generates Excel and CSV reports with color highlighting
"""

import pandas as pd
from typing import Dict, Any, List, Optional
from utils import get_logger, ReportMetadata, FileUtils
from data_compare import ComparisonResult
from datetime import datetime
import os


class ReportGenerator:
    """Generates comparison reports in Excel and CSV formats"""
    
    # Color codes for highlighting (RGB format for openpyxl)
    COLORS = {
        'red': 'FFFF0000',        # Mismatched values
        'yellow': 'FFFFFF00',     # Missing records
        'green': 'FF00B050',       # Matched records
        'light_red': 'FFFFE6E6',  # Light red background
        'light_yellow': 'FFFFFF99',  # Light yellow background
        'light_green': 'FFE2EFDA' # Light green background
    }
    
    def __init__(self, output_dir: str = "./reports"):
        self.output_dir = output_dir
        self.logger = get_logger('ReportGenerator')
        FileUtils.ensure_directory_exists(output_dir)
        FileUtils.ensure_directory_exists(os.path.join(output_dir, 'excel'))
        FileUtils.ensure_directory_exists(os.path.join(output_dir, 'csv'))
    
    def generate_excel_report(self, comparison_result: ComparisonResult,
                             source_df: pd.DataFrame,
                             target_df: pd.DataFrame,
                             source_name: str = "Source",
                             target_name: str = "Target",
                             filename: Optional[str] = None,
                             primary_keys: Optional[List[str]] = None) -> str:
        """Generate comprehensive Excel report with multiple sheets"""
        
        if filename is None:
            filename = FileUtils.generate_report_filename("comparison_report", "xlsx")
        
        file_path = os.path.join(self.output_dir, "excel", filename)
        
        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Sheet 1: Summary
                self._write_summary_sheet(comparison_result, source_name, target_name,
                                        writer, primary_keys)
                
                # Sheet 2: Matched Records
                matched_df = self._get_matched_records(source_df, target_df,
                                                      comparison_result, primary_keys)
                matched_df.to_excel(writer, sheet_name='Matched_Records', index=False)
                self._highlight_sheet(writer, 'Matched_Records', 'green')
                
                # Sheet 3: Mismatched Records
                if comparison_result.mismatch_details:
                    mismatch_df = self._create_mismatch_dataframe(
                        comparison_result.mismatch_details, source_df, target_df,
                        primary_keys
                    )
                    mismatch_df.to_excel(writer, sheet_name='Mismatched_Records', index=False)
                    self._highlight_sheet(writer, 'Mismatched_Records', 'red')
                
                # Sheet 4: Missing in Target
                if comparison_result.missing_record_keys:
                    missing_df = self._get_missing_records(source_df, comparison_result,
                                                          primary_keys)
                    missing_df.to_excel(writer, sheet_name='Missing_in_Target', index=False)
                    self._highlight_sheet(writer, 'Missing_in_Target', 'yellow')
                
                # Sheet 5: Extra in Target
                if comparison_result.extra_record_keys:
                    extra_df = self._get_extra_records(target_df, comparison_result,
                                                      primary_keys)
                    extra_df.to_excel(writer, sheet_name='Extra_in_Target', index=False)
                    self._highlight_sheet(writer, 'Extra_in_Target', 'yellow')
            
            self.logger.info(f"Excel report generated: {file_path}")
            return file_path
        
        except Exception as e:
            error_msg = f"Error generating Excel report: {str(e)}"
            self.logger.error(error_msg)
            raise Exception(error_msg)
    
    def generate_csv_summary(self, comparison_result: ComparisonResult,
                            source_name: str = "Source",
                            target_name: str = "Target",
                            metadata: Optional[Dict[str, Any]] = None,
                            filename: Optional[str] = None) -> str:
        """Generate CSV summary report"""
        
        if filename is None:
            filename = FileUtils.generate_report_filename("comparison_summary", "csv")
        
        file_path = os.path.join(self.output_dir, "csv", filename)
        
        try:
            summary_data = comparison_result.to_dict()
            summary_data['source_name'] = source_name
            summary_data['target_name'] = target_name
            summary_data['execution_timestamp'] = datetime.now().isoformat()
            
            if metadata:
                summary_data.update(metadata)
            
            df = pd.DataFrame([summary_data])
            df.to_csv(file_path, index=False)
            
            self.logger.info(f"CSV summary report generated: {file_path}")
            return file_path
        
        except Exception as e:
            error_msg = f"Error generating CSV report: {str(e)}"
            self.logger.error(error_msg)
            raise Exception(error_msg)
    
    def _write_summary_sheet(self, result: ComparisonResult, source_name: str,
                            target_name: str, writer, primary_keys: Optional[List[str]]):
        """Write summary sheet to Excel"""
        summary_data = {
            'Metric': [
                'Source Name',
                'Target Name',
                'Execution Time',
                'Total Source Records',
                'Total Target Records',
                'Matched Records',
                'Mismatched Records',
                'Missing Records',
                'Extra Records',
                'Overall Status'
            ],
            'Value': [
                source_name,
                target_name,
                datetime.now().isoformat(),
                result.total_source_records,
                result.total_target_records,
                result.matched_records,
                result.mismatched_records,
                result.missing_records,
                result.extra_records,
                result.status
            ]
        }
        
        if primary_keys:
            summary_data['Metric'].append('Primary Keys')
            summary_data['Value'].append(', '.join(primary_keys))
        
        df = pd.DataFrame(summary_data)
        df.to_excel(writer, sheet_name='Summary', index=False)
    
    def _get_matched_records(self, source_df: pd.DataFrame, target_df: pd.DataFrame,
                            result: ComparisonResult, 
                            primary_keys: Optional[List[str]]) -> pd.DataFrame:
        """Extract matched records"""
        if not primary_keys or result.matched_records == 0:
            return pd.DataFrame()
        
        source_keys = set(map(tuple, source_df[primary_keys].values))
        target_keys = set(map(tuple, target_df[primary_keys].values))
        matched_keys = source_keys & target_keys
        
        if isinstance(primary_keys[0], str):
            # Single key
            matched_df = source_df[source_df[primary_keys[0]].isin([k[0] for k in matched_keys])]
        else:
            matched_df = source_df.copy()
        
        return matched_df
    
    def _create_mismatch_dataframe(self, mismatch_details: List[Dict],
                                  source_df: pd.DataFrame, target_df: pd.DataFrame,
                                  primary_keys: Optional[List[str]]) -> pd.DataFrame:
        """Create DataFrame from mismatch details"""
        records = []
        
        for mismatch in mismatch_details[:100]:  # Limit to 100 samples
            key = mismatch['key']
            for col_mismatch in mismatch['column_mismatches']:
                records.append({
                    'Primary_Key': str(key),
                    'Column': col_mismatch['column'],
                    'Source_Value': col_mismatch['source_value'],
                    'Target_Value': col_mismatch['target_value']
                })
        
        return pd.DataFrame(records) if records else pd.DataFrame()
    
    def _get_missing_records(self, source_df: pd.DataFrame,
                            result: ComparisonResult,
                            primary_keys: Optional[List[str]]) -> pd.DataFrame:
        """Extract records missing in target"""
        if not primary_keys or not result.missing_record_keys:
            return pd.DataFrame()
        
        # Filter source_df to only missing records
        if len(primary_keys) == 1:
            missing_df = source_df[source_df[primary_keys[0]].isin(
                [k[0] for k in result.missing_record_keys]
            )]
        else:
            # Multiple keys - filter by combination
            key_df = pd.DataFrame(result.missing_record_keys, columns=primary_keys)
            missing_df = source_df.merge(key_df, on=primary_keys, how='inner')
        
        return missing_df
    
    def _get_extra_records(self, target_df: pd.DataFrame,
                          result: ComparisonResult,
                          primary_keys: Optional[List[str]]) -> pd.DataFrame:
        """Extract records extra in target"""
        if not primary_keys or not result.extra_record_keys:
            return pd.DataFrame()
        
        # Filter target_df to only extra records
        if len(primary_keys) == 1:
            extra_df = target_df[target_df[primary_keys[0]].isin(
                [k[0] for k in result.extra_record_keys]
            )]
        else:
            key_df = pd.DataFrame(result.extra_record_keys, columns=primary_keys)
            extra_df = target_df.merge(key_df, on=primary_keys, how='inner')
        
        return extra_df
    
    def _highlight_sheet(self, writer, sheet_name: str, color_type: str):
        """Apply highlighting to sheet"""
        try:
            from openpyxl.styles import PatternFill
            
            workbook = writer.book
            worksheet = workbook[sheet_name]
            
            # Determine color
            color = self.COLORS.get(color_type, self.COLORS['light_green'])
            fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            
            # Apply to all data rows (skip header)
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row,
                                          min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.fill = fill
        
        except ImportError:
            self.logger.warning("openpyxl not available for highlighting")
        except Exception as e:
            self.logger.warning(f"Error applying highlighting: {str(e)}")


# Utility function to create generator
def create_report_generator(output_dir: str = "./reports") -> ReportGenerator:
    """Factory function to create ReportGenerator"""
    return ReportGenerator(output_dir)
